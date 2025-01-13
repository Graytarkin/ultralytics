import os
import cv2
import glob
from openpyxl import Workbook
import numpy as np

## Good !!

files = glob.glob("C:/yolo/AI-Pred-Data/step5" + '/*.jpeg', recursive=True)
dir_path = "C:/yolo/AI-Pred-Data/step5/"

wb = Workbook()
ws = wb.active
exLine = 2

for file in files:
    filename = os.path.basename(file)
    print('File:' + filename)
    img = cv2.imread(file)

    filenamedPos = filename.rfind("-")
    filenameNM = filename[:filenamedPos]

    gimg = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # 収縮でノイズを消す
    kernele = np.ones((5, 5), np.uint8)
    erosion = cv2.erode(gimg, kernele, iterations=1)
    # 膨張で文字領域を拡大
    kerneld = np.ones((5, 5), np.uint8)
    dgimg = cv2.dilate(erosion, kerneld, iterations=1)
 
    size = (50, 50)  # 分割後の大きさ
    rows = int(np.ceil(dgimg.shape[0] / size[0]))  # 行数
    cols = int(np.ceil(dgimg.shape[1] / size[1]))  # 列数

    print(filenameNM, rows, cols)

    mrows = int(np.ceil(dgimg.shape[0] / size[0]))  # 行数
    mcols = int(np.ceil(dgimg.shape[1] / size[1]))  # 列数

    msize = [mrows, mcols]
    print(msize)

    chunks = []
    chunks.append(msize)
    for mrow_img in np.array_split(dgimg, mrows, axis=0):
        for chunk in np.array_split(mrow_img, mcols, axis=1):
            chunks.append(chunk)
    print(len(chunks))

    popped_item = chunks.pop(0)
    Rows, Cols = popped_item
    print(Rows, Cols,  len(chunks))
    colsImage = []
    rowImage = []  
    j = 0
    for imga in chunks:
#        gimg = cv2.cvtColor(imga, cv2.COLOR_BGR2GRAY)

        ret, imgT = cv2.threshold(imga, 0, 255, cv2.THRESH_OTSU)
        std = np.std(imga)
        median = np.median(imga)
        if abs(median - ret) < 5 and std < 10:
            ret2, imgT = cv2.threshold(imga, 100, 255, cv2.THRESH_BINARY)
        rowImage.append(imgT)
        j = j + 1
        if j == Cols:
            image_v = cv2.hconcat(rowImage) 
            rowImage = [] 
            j = 0  
            colsImage.append(image_v)  
    image_H = cv2.vconcat(colsImage)


    ohight, owidth = image_H.shape
    result = cv2.resize(image_H, (owidth * 1, ohight * 2)) 
    imgfroct = cv2.cvtColor(result, cv2.COLOR_GRAY2BGR)
    # グレースケール
    gray = result
    #gray = cv2.cvtColor(image_H, cv2.COLOR_BGR2GRAY)

    # 判定画像
    minlength = 10
    gap = 0
    kernelJ = cv2.getStructuringElement(cv2.MORPH_CROSS, (3, 3))
    dst = cv2.dilate(gray, kernelJ)

    imgMB = cv2.medianBlur(gray, 5)
    #    judge_imgN = cv2.adaptiveThreshold(imgMB, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    #    judge_imgN = cv2.adaptiveThreshold(imgMB, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 7, 2)
    judge_imgN = cv2.adaptiveThreshold(imgMB, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    #    imgVV = cv2.cvtColor(judge_imgN, cv2.COLOR_BGR2GRAY)
    judge_img = cv2.bitwise_not(judge_imgN)

    lines = cv2.HoughLinesP(judge_img, rho=1, theta=np.pi / 360, threshold=100, minLineLength=100, maxLineGap=5)
    print(len(lines))
    no_lines_img = imgfroct
    for line in lines:
        x1, y1, x2, y2 = line[0]
        # 線を消す(白で線を引く)
        if abs(y1 -y2) < 50 or abs(x1 -x2) / abs(y1 -y2) < 0.25:
            no_lines_img = cv2.line(imgfroct, (x1,y1), (x2,y2), (0,255,0), 10)
#        cv2.imwrite(dir_path + "W" + filename + ".jpeg", no_lines_img)
    for line2 in lines:
        x1, y1, x2, y2 = line2[0]
        # 線を消す(白で線を引く)
        if abs(x1 -x2) < 50:
            cv2.line(no_lines_img, (x1,y1), (x2,y2), (0,255,0), 5)

#    cv2.imwrite(dir_path + "W" + filename + ".jpeg", no_lines_img)

#    gray2 = no_lines_img
    gray2 = cv2.cvtColor(no_lines_img, cv2.COLOR_BGR2GRAY)
    # 閾値の設定
    threshold = 100
    # 二値化(閾値100を超えた画素を255にする。)
    # ret, img_thresh = cv2.threshold(gray2, threshold, 255, cv2.THRESH_BINARY)

    # 二値化(大津)
    ret2, img_otsu = cv2.threshold(gray2, 0, 255, cv2.THRESH_OTSU)

#    cv2.imwrite(dir_path + "X" + filename + ".jpeg", img_otsu)
    img_otsuN = 255 - img_otsu
    # 収縮でノイズを消す
    kernele = np.ones((3, 3), np.uint8)
    erosion = cv2.erode(img_otsuN, kernele, iterations=1)
#    cv2.imwrite(dir_path + "Y" + filename + ".jpeg", erosion)
    # 膨張で文字領域を拡大
    kerneld = np.ones((5, 50), np.uint8)
    dilation = cv2.dilate(erosion, kerneld, iterations=1)
#    cv2.imwrite(dir_path + "Z" + filename + ".jpeg", dilation)
    for line in lines:
        x1, y1, x2, y2 = line[0]
        # 線を消す(白で線を引く)
        if abs(y1 -y2) < 50 or abs(x1 -x2) / abs(y1 -y2) < 0.25:
            no_lines_img = cv2.line(imgfroct, (x1,y1), (x2,y2), (255,255,255), 10)
#        cv2.imwrite(dir_path + "W" + filename + ".jpeg", no_lines_img)
    for line2 in lines:
        x1, y1, x2, y2 = line2[0]
        # 線を消す(白で線を引く)
        if abs(x1 -x2) < 50:
            cv2.line(no_lines_img, (x1,y1), (x2,y2), (255,255,255), 3)



    contours, hierarchy = cv2.findContours(dilation, cv2.RETR_LIST, cv2.CHAIN_APPROX_TC89_L1, )

    # 画像表示用に入力画像をカラーデータに変換する
#    img_disp = cv2.cvtColor(no_lines_img, cv2.COLOR_GRAY2BGR)
    img_disp = img
    ihight, iwidth, icolor = img_disp.shape
    disp2 = cv2.resize(img_disp, (iwidth * 1, ihight * 2)) 

    bigRect = []
    for i, contour in enumerate(contours):
# 輪郭を描画
        area = cv2.contourArea(contour)
        if 9000 < area < 500000:

            # 傾いた外接する矩形領域
            rect = cv2.minAreaRect(contour)
            box = cv2.boxPoints(rect)
            box = np.intp(box)
            pa,pb,pc,pd = box
            pax, pay = pa
            pbx, pby = pb
            pcx, pcy = pc
            pdx, pdy = pd

#            bigRect.append(box)
            cv2.drawContours(disp2, [box], 0, (255, 0, 0), 4)
            ws.cell(row=exLine, column=1).value = filenameNM          
            ws.cell(row=exLine, column=2).value = int(pax)
            ws.cell(row=exLine, column=3).value = int(pay / 2)
            ws.cell(row=exLine, column=4).value = int(pbx)
            ws.cell(row=exLine, column=5).value = int(pby / 2) 
            ws.cell(row=exLine, column=6).value = int(pcx)
            ws.cell(row=exLine, column=7).value = int(pcy / 2)
            ws.cell(row=exLine, column=8).value = int(pdx)
            ws.cell(row=exLine, column=9).value = int(pdy / 2)   
            ws.cell(row=exLine, column=10).value = int(exLine)  
            exLine += 1

        if i > 500000:
            break
    fohight, fowidth, focolor = disp2.shape
    efhight = int(fohight / 2)
    resultF = cv2.resize(disp2, (fowidth, efhight)) 
    cv2.imwrite(dir_path + "test" + filename + ".jpeg", resultF)


wb.save(dir_path + "bigRect" + ".xlsx")    