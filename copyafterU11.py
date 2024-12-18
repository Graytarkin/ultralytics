import csv
import os
import glob
import hdbscan
import heapq
import numpy as np
from io import BytesIO
from sklearn.neighbors import KernelDensity
from scipy.signal import argrelextrema
from PIL import Image as Imgpl
# from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as Imgxl

dirOfPredx = "C:/yolo/AI-Pred-Data/step6/"

# 各クラスタリングアルゴリズムの設定
algorithms = [3,4,5,6,7,8]

with open("C:/yolo/AI-Pred-Data/step5/predictions.csv", encoding="utf-8") as file:
#with open("C:/yolo/ultralytics/runs/detect/predict/predictions.csv", encoding="utf-8") as file:
    lst = list(csv.reader(file))

saveF = ""
excelWRl = []
lcnt = 0
for listl in lst:
    lcnt = lcnt + 1
    dotPos = listl[0].rfind(".")
    listltype = listl[0][dotPos + 1:]
    listlname = listl[0][:dotPos]
#    print("listname",listlname)
    if "PMX" in listlname:
        listltype = "png"
    if listl[0][:dotPos] != saveF:
        saveF = listl[0][:dotPos]
        textCnt = 0
        with open("C:/yolo/ultralytics/runs/detect/predict/labels/" + saveF + ".txt", "r", encoding="utf-8") as f:
            listtxt = f.readlines()
            listtxtS = []
            for listelm in listtxt:
                valofelm = listelm.split()
                listtxtS.append(valofelm)
#            print(len(listtxt),listtxtS)
    if listltype == "png":
        posfnameE = listl[0].rfind("PMX")
        fname = listl[0][:posfnameE]
    else:
        posfnameE = listl[0].rfind("-N")
        fname = listl[0][:posfnameE]

#YOLOv11
    elmm = [listl[0],listl[2],listtxtS[textCnt][5],listtxtS[textCnt][0],listtxtS[textCnt][1],listtxtS[textCnt][2],listtxtS[textCnt][3],listtxtS[textCnt][4],fname]
#YOLOv5                 
#    elmm = [listl[0],listl[1],listl[2],listtxtS[textCnt][0],listtxtS[textCnt][1],listtxtS[textCnt][2],listtxtS[textCnt][3],listtxtS[textCnt][4],fname]
    textCnt +=1
    excelWRl.append(elmm)
#print(excelWRl)

types = ('jpg','jpeg','png')
filespg = []
for t in types:
    filespg += glob.glob("C:/yolo/ultralytics/runs/detect/predict/*." + t, recursive=True)
print(len(filespg))

for file in filespg:
    filename = os.path.basename(file)
    dotPos = filename[0].rfind(".")
#    listltype = listl[0][dotPos + 1:]
    listlname = filename[0][:dotPos]
#    print("listname2",listlname)
    if "PMX" in listlname:
        listltype = "png"
    else:
        listltype = "jpg"       
    listlname = listl[0][:dotPos]
#    print("listname3",listlname)
    if "PMX" in listlname:
        posfnameE = listl[0].rfind("PMX")
    else:
        posfnameE = listl[0].rfind("-N")
    if not(any(filename in row for row in excelWRl)):
        elmm = [filename,'','0','0','0','0','0','0',fname]
        textCnt +=1
        excelWRl.append(elmm)

sortedexcelWRl = sorted(excelWRl, key=lambda x:(x[0], x[4]))

exName = ""
exPosR = 0
exPosC = 0

wb = Workbook()
ws = wb.active

fnamesv = ""
mkexposl = []

for dataelm in sortedexcelWRl:
    exName = dataelm
    exPosR = exPosR + 1
#   
    posDot = exName[0].rfind(".")
#    filetype = exName[0][posDot + 1:]
    filename = exName[0][:posDot]
    if "PMX" in filename:
        filetype = "png"
        posfnameE = listl[0].rfind("PMX")
    else:
        filetype = "jpg"
        posfnameE = listl[0].rfind("-N")
#    print(filetype)
    if filetype == "png":
        posRCC = exName[0].rfind("C")
        posRCR = exName[0].rfind("-R")
        posSX = exName[0].rfind("-X")
        posSY = exName[0].rfind("Y")
        posN = exName[0].rfind("-N")
        posMX = exName[0].rfind("PMX")
        posMY = exName[0].rfind("MY")

        fname = exName[0][:posMX]
        filepathname = exName[0][:posMX + 1]
        fMX = int(exName[0][posMX + 3: posMY])
        fMY = int(exName[0][posMY + 2: posN])
        fNo = int(exName[0][posN + 2: posSX])
        fwx = int(exName[0][posSX + 2: posSY])
        fhy = int(exName[0][posSY + 1: posRCR])
        frr = int(exName[0][posRCR + 2: posRCC])
        fcc = int(exName[0][posRCC + 1: posDot])

    else:
        posRCC = exName[0].rfind("C")
        posRCR = exName[0].rfind("-R")
        posSX = exName[0].rfind("-X")
        posSY = exName[0].rfind("Y")
        posN = exName[0].rfind("-N")

        fname = exName[0][:posN]
        filepathname = exName[0][:posN + 1]
        fMX = 0
        fMY = 0
        fNo = int(exName[0][posN + 2: posSX])
        fwx = int(exName[0][posSX + 2: posSY])
        fhy = int(exName[0][posSY + 1: posRCR])
        frr = int(exName[0][posRCR + 2: posRCC])
        fcc = int(exName[0][posRCC + 1: posDot])

#
    ws.cell(row=exPosR, column=1).value = str(exName[0])
    ws.cell(row=exPosR, column=2).value = filetype    
    ws.cell(row=exPosR, column=3).value = str(exName[1])
    lLetter = str(exName[1])
    ws.cell(row=exPosR, column=4).value = str(exName[2])
    lLpercent = int(float(exName[2]) * 100)
    ws.cell(row=exPosR, column=5).value = str(exName[3]) 
    lLclassno = int(exName[3])   
    ws.cell(row=exPosR, column=6).value = str(exName[4])
    ws.cell(row=exPosR, column=7).value = str(exName[5])
    ws.cell(row=exPosR, column=8).value = str(exName[6])
    ws.cell(row=exPosR, column=9).value = str(exName[7])
    ws.cell(row=exPosR, column=11).value = filepathname
    ws.cell(row=exPosR, column=11).value = fname
    ws.cell(row=exPosR, column=12).value = fMX
    ws.cell(row=exPosR, column=13).value = fMY
    ws.cell(row=exPosR, column=14).value = fNo
    ws.cell(row=exPosR, column=15).value = fwx
    ws.cell(row=exPosR, column=16).value = fhy
    ws.cell(row=exPosR, column=17).value = frr
    ws.cell(row=exPosR, column=18).value = fcc      

    if filepathname != fnamesv:
        fnamesv = filepathname
        addwcol = 0
        addhrow = 0
        fwcolsv = fwx
        fhrowsv = fhy
        rowaddsv = -1
        fnosv = -1
    rowadd = fNo // fcc
    if rowadd != rowaddsv:
        addwcol = 0
        rowaddsv = rowadd
        addhrow = fhy + addhrow      
    if fNo != fnosv:
        fnosv = fNo
        addwcol = fwx + addwcol
    
    fforgY = fMY + addhrow - fhrowsv
    fforgX = fMX + addwcol - fwcolsv
    ws.cell(row=exPosR, column=20).value = fforgY
    ws.cell(row=exPosR, column=21).value = fforgX 

#    lFwx = int(float(exName[4]) * float(fwx) * 1)+ int(float(exName[6]) * float(fwx) * 0.5)
    lFwx = int(float(exName[4]) * float(fwx) * 1) - int(float(exName[6]) * float(fwx) * 0.5)
    ws.cell(row=exPosR, column=22).value = lFwx
    lFhy = int(float(exName[5]) * float(fhy) * 1) 
    ws.cell(row=exPosR, column=23).value = lFhy 
    areaw = int(float(exName[6]) * float(fwx) * float(exName[7]) * float(fhy) * 0.1)
    ws.cell(row=exPosR, column=24).value = areaw
            
    ttlFhy = fforgY + lFhy
    ttlFwx = fforgX + lFwx
#    ttlFhy2 = ttlFhy * 2
#
    ws.cell(row=exPosR, column=25).value = ttlFwx 
    ws.cell(row=exPosR, column=26).value = ttlFhy
    leftend = int((float(exName[4]) - (float(exName[6]) * 0.5)) *  float(fwx)) + fforgX
    rightend = int((float(exName[4]) + (float(exName[6]) * 0.5)) *  float(fwx)) + fforgX
#               fn:0     ft:1   char:2    cls:3      rate:4     tx:5    ty:6    le:7      re:8    aw:9
    mkexpos = [fname, filetype, lLetter, lLclassno, lLpercent, ttlFwx, ttlFhy, leftend, rightend, areaw]
    if lLpercent > 0:
        mkexposl.append(mkexpos)

wb.save(dirOfPredx + "exName" + ".xlsx")     

sdnposl = sorted(mkexposl, key=lambda x:(x[0],x[6])) 

clstsv = []
clstsvl = []
score2 = []
datafilesv = ""
for i in range(0,len(sdnposl)):
    datafile = sdnposl[i][0]
    if datafilesv != datafile:
        if datafilesv != "":
            column = [row[0] for row in score2]
            row = [row[1] for row in score2]
            score3 = np.array(score2)

            X = score3
            cntminsv = 999
            for algo in algorithms:
                if len(X)> 1:
                    cluster = hdbscan.HDBSCAN(min_cluster_size=algo).fit_predict(X)
                    cntmin = np.sum(cluster < 0)
                    if cntmin <= cntminsv:
                        cntminsv = cntmin
                        clstsv = cluster
            clstsvl.append(clstsv)    
            score2 = []           
        datafilesv = datafile         
    cellx = sdnposl[i][5] / 70
    celly = sdnposl[i][6] * 1
    scoreelm = [cellx, celly]
    score2.append(scoreelm)

column = [row[0] for row in score2]
row = [row[1] for row in score2]
score3 = np.array(score2)

X = score3
cntminsv = 999
for algo in algorithms:
    if len(X)> 1:
        cluster = hdbscan.HDBSCAN(min_cluster_size=algo).fit_predict(X)
        cntmin = np.sum(cluster < 0)
        if cntmin <= cntminsv:
            cntminsv = cntmin
            clstsv = cluster
clstsvl.append(clstsv) 

addi = 0
addj = 0
sdnpcls = []
sdnpclsmin = []
for i in range(0,len(sdnposl)):
    if addi >= len(clstsvl[addj]):
        addj = addj + 1
        addi = 0
    lcls = clstsvl[addj][addi]
    addi = addi + 1
    sdnpclsElm = [sdnposl[i][0],sdnposl[i][1],sdnposl[i][2],sdnposl[i][3],sdnposl[i][4],
                  sdnposl[i][5],sdnposl[i][6],sdnposl[i][7],sdnposl[i][8],sdnposl[i][9],0,lcls]
    if lcls < 0:
        sdnpclsmin.append(sdnpclsElm)
    else:
        sdnpcls.append(sdnpclsElm)

sdnclss = sorted(sdnpcls, key=lambda x:(x[0],x[6])) 
sdnclsmins = sorted(sdnpclsmin, key=lambda x:(x[0],x[6])) 


for imin in range(0,len(sdnclsmins)):
    flgfound = 0
    tempappc = []
    for icls in range(0,len(sdnclss)):
        if sdnclsmins[imin][0] == sdnclss[icls][0]:
            appclsElm = [sdnclss[icls][6], sdnclss[icls][11]]
            tempappc.append(appclsElm) 
    tempappci = [row[0] for row in tempappc]
    idx = np.abs(np.asarray(tempappci) - int(sdnclsmins[imin][6])).argmin()
    sdnclsmins[imin][11] = tempappc[idx][1]
    sdnclsmins[imin][10] = -1 
 
sdnclss.extend(sdnclsmins)

serai = sorted(sdnclss, key=lambda x:(x[0],x[11],x[5]))    

derai = []
cherai = []
meanscore = []
datafckey = ""
######
for i in range(0,len(serai)):
    datafc = serai[i][0] + str(serai[i][11])
    if datafckey != datafc:
        if datafckey != "":
            column = [row[0] for row in meanscore]
            row = [row[1] for row in meanscore]
            scorenp = np.array(meanscore)

# 各クラスタリングアルゴリズムの設定
            maxclsl = []
            for ai in range(2,7):
                if len(scorenp)> 1:
                    cluster = hdbscan.HDBSCAN(min_cluster_size=ai).fit_predict(scorenp)
                    cntmin = np.sum(cluster < 0)
                    maxclsElm = [cntmin,max(cluster),cluster,ai]
                    maxclsl.append(maxclsElm)
            maxclsls = sorted(maxclsl, key=lambda x:(x[3]),reverse=True)
            algomin = 99
            for si in range(0,len(maxclsls)):
                if maxclsls[si][0] == 0:
                    if algomin > maxclsls[si][1]:
                        if maxclsls[si][1] > 1:
                            algomin = maxclsls[si][1]
                            selectc = maxclsls[si][2]
            if algomin == 99:
                outmin = 99
                j = 0
                for ix in range(0,len(maxclsls)):
                    if maxclsls[ix][0] < outmin:
                        outmin = maxclsls[ix][0]
                        j = ix
                if outmin != 99:
                    selectc = maxclsls[j][2]

        for di in range(0,len(derai)): 
            ncElm = derai[di]
            if selectc[di] < 0:
                ncElm[10] = 88
            else:
                ncElm[10] = selectc[di]
            cherai.append(ncElm)
        datafckey = datafc
        derai = []
        meanscore = []
    scoreelm = [serai[i][5], serai[i][6]]
    meanscore.append(scoreelm)
    derai.append(serai[i])

column = [row[0] for row in meanscore]
row = [row[1] for row in meanscore]
scorenp = np.array(meanscore)

maxclsl = []
for ai in range(2,7):
    if len(scorenp)> 1:
        cluster = hdbscan.HDBSCAN(min_cluster_size=ai).fit_predict(scorenp)
        cntmin = np.sum(cluster < 0)
        maxclsElm = [cntmin,max(cluster),cluster,ai]
        maxclsl.append(maxclsElm)
    maxclsls = sorted(maxclsl, key=lambda x:(x[3]),reverse=True)
    algomin = 99
    for i in range(0,len(maxclsls)):
        if maxclsls[i][0] == 0:
            if algomin > maxclsls[i][1]:
                if maxclsls[i][1] > 1:
                    algomin = maxclsls[i][1]
                    selectc = maxclsls[i][2]
    if algomin == 99:
        outmin = 99
        j = 0
        for i in range(0,len(maxclsls)):
            if maxclsls[i][0] < outmin:
                outmin = maxclsls[i][0]
                j = i
            if outmin != 99:
                selectc = maxclsls[j][2]

cntMinus = 0
for i in range(0,len(derai)): 
    ncElm = derai[i]
    if selectc[i] < 0:
        ncElm[10] = 88
    else:
        ncElm[10] = selectc[i]
    cherai.append(ncElm)

serai = sorted(cherai, key=lambda x:(x[0],x[11],x[10],x[5])) 


for i in reversed(range(1, len(serai))):
    if abs(serai[i - 1][5] - serai[i][5]) < 5 and abs(serai[i - 1][6] - serai[i][6]) < 5 and \
       abs(serai[i - 1][7] - serai[i][7]) < 5 and abs(serai[i - 1][8] - serai[i][8]) < 5 and \
       (serai[i - 1][0] == serai[i][0]) and (serai[i - 1][3] == serai[i][3]) and \
       (serai[i - 1][11] == serai[i][11]) and (serai[i - 1][10] == serai[i][10]):
        if (serai[i - 1][4] > serai[i][4]):
            perai = serai.pop(i)
        else:
            perai = serai.pop(i -1)
        j = i + 2

temperai = []

Groupnum = 1
terai = []
teraiElm = [serai[0][0],serai[0][1],serai[0][2],serai[0][3],serai[0][4],serai[0][5],
            serai[0][6],serai[0][7],serai[0][8],serai[0][9],serai[0][10],serai[0][11],0,0,0]
terai.append(teraiElm)
for i in range(1, len(serai)):
    if ((serai[i][0]!=serai[i-1][0]) or (serai[i][11]!=serai[i-1][11]) or (serai[i][10]!=serai[i-1][10])) or \
       (abs(serai[i][5]-serai[i-1][5]) > 10 and abs(serai[i][7]-serai[i-1][7]) > 10 and abs(serai[i][8]-serai[i-1][8]) > 10) or \
        abs(serai[i][6]-serai[i-1][6]) > 30                                                                                                                                                                                                  :
        x78dif = 0
    else:
        x78dif = serai[i][7]-serai[i -1][8]
    
    if x78dif <  -10:
        terai[i - 1][13] = Groupnum 
        teraiElm = [serai[i][0],serai[i][1],serai[i][2],serai[i][3],serai[i][4],serai[i][5],
            serai[i][6],serai[i][7],serai[i][8],serai[i][9],serai[i][10],serai[i][11],x78dif,Groupnum,0]
        terai.append(teraiElm)
    else:
        Groupnum =  Groupnum + 1
        teraiElm = [serai[i][0],serai[i][1],serai[i][2],serai[i][3],serai[i][4],serai[i][5],
            serai[i][6],serai[i][7],serai[i][8],serai[i][9],serai[i][10],serai[i][11],x78dif,0,0] 
        terai.append(teraiElm)


svGroup = 0
tterai = []
ttterai = []
erakunai = []
for i in range(0,len(terai)):
    if terai[i][13] > 0:
        if svGroup != terai[i][13]:
            if svGroup != 0:
                if len(tterai) > 0:
                    ttterai.append(tterai)
            svGroup = terai[i][13]        
            tterai = []
            tterai.append(terai[i])
        else:
            tterai.append(terai[i])
    else:
        kunaiElm = [terai[i][0],terai[i][1],terai[i][2],terai[i][3],terai[i][4],terai[i][5],terai[i][6],terai[i][7],
                    terai[i][8],terai[i][9],terai[i][10],terai[i][11],terai[i][12],terai[i][13],terai[i][14]]
        erakunai.append(kunaiElm)
        if len(tterai) > 0:
            ttterai.append(tterai)
            tterai = []
if len(tterai) > 0:
    ttterai.append(tterai)        


def DuplDelete(listx):
    steraid = []
    steraind = []
    sterai = listx
    for j in reversed(range(1,len(sterai))):
        absFour = []
        absFour.append(abs(sterai[j][5] - sterai[0][5]))
        absFour.append(abs(sterai[j][6] - sterai[0][6]))
        absFour.append(abs(sterai[j][7] - sterai[0][7]))
        absFour.append(abs(sterai[j][8] - sterai[0][8]))
        mchcnt = sum(x < 7 for x in absFour)     
        absFourS = sorted(absFour)
        s3sum = sum(heapq.nsmallest(3, absFourS))
        s2sum = sum(heapq.nsmallest(2, absFourS))
        s1min = min(absFourS)
        ratewidth = float(int(sterai[j][9]))/ float(int(sterai[0][9]))
        if ratewidth > 1.22 or ratewidth < 0.82:
            ratelbl = 3
        elif ratewidth > 1.213 or ratewidth < 0.825:
            ratelbl = 2
        elif ratewidth > 1.121 or ratewidth < 0.892:
            ratelbl = 1
        else:       
            ratelbl = 0
        jpegpngF = 0
        if sterai[j][1] != sterai[0][1]:
            jpegpngF = 1

        flgPop = 0
        if sterai[j][3] == sterai[0][3]:
            if mchcnt > 2:
                steraieLm14 = mchcnt * 10
                flgPop = 1
            elif mchcnt > 1:
                if s2sum < 6 or ratelbl > 2:
                    steraieLm14 = mchcnt * 10 + ratelbl
                    flgPop = 1
                else:
                    if s2sum < 8 and jpegpngF > 0:
                        steraieLm14 = 510 + ratelbl
                        flgPop = 1
                    else:
                        steraieLm14 = 910 + ratelbl
                        flgPop = 2
            elif mchcnt > 0:
                if s1min < 5 and ratelbl > 0:
                    steraieLm14 = mchcnt * 10 + ratelbl
                    flgPop = 1
                else:
                    steraieLm14 = 920 + ratelbl
                    flgPop = 2
            else:
                steraieLm14 = 930 + ratelbl
                flgPop = 2                
        else:
            if mchcnt > 3:
                steraieLm14 = mchcnt * 10 + ratelbl
                flgPop = 1
            elif mchcnt > 2:
                if s3sum < 6 or ratelbl > 0:                
                    steraieLm14 = mchcnt * 10 + ratelbl
                    flgPop = 1
                else:
                    if s2sum < 3 and jpegpngF > 0:
                        steraieLm14 = 520 + ratelbl
                        flgPop = 1
                    else:
                        steraieLm14 = 940 + ratelbl
                        flgPop = 2
            elif mchcnt > 1:
                if ratelbl > 2:
                    steraieLm14 = mchcnt * 10 + ratelbl
                    flgPop = 1
                elif s2sum < 5 and jpegpngF > 0:
                    steraieLm14 = 530 + ratelbl
                    flgPop = 1
                else:
                    steraieLm14 = 950 + ratelbl
                    flgPop = 2
            elif mchcnt > 0:
                if s1min < 2 and ratelbl > 2: 
                    steraieLm14 = mchcnt * 10 + ratelbl
                    flgPop = 1                      
                else:
                    steraieLm14 = 960 + ratelbl
                    flgPop = 2
            else:
                steraieLm14 = 970 + ratelbl
                flgPop = 2      
        kElm = [sterai[j][0],sterai[j][1],sterai[j][2],sterai[j][3],sterai[j][4],sterai[j][5],sterai[j][6],
               sterai[j][7],sterai[j][8],sterai[j][9],sterai[j][10],sterai[j][11],sterai[j][12],sterai[j][13],steraieLm14]

        if flgPop == 1 and sterai[j][4] <= sterai[0][4]:
            pop = sterai.pop(j)
            steraid.append(kElm)
        else:
            pop = sterai.pop(j)
            steraind.append(kElm)            
    return sterai,steraid,steraind

tderai = []
for i in range(0,len(ttterai)):
    stterai = sorted(ttterai[i], key=lambda x:(x[4]),reverse=True)
    stterai[0][14] = 999
    l1, l2, l3 = DuplDelete(stterai)
    if len(l3) > 0:
        if len(l2) > 0:
            erakunai.extend(l2)
        l1.extend(l3)
#        print(l1)
        tderai = sorted(l1, key=lambda x:(x[14]))
        l1, l2, l3 = DuplDelete(tderai)
        erakunai.extend(l1)
        erakunai.extend(l2)
        erakunai.extend(l3)
    else:
        erakunai.extend(l2)        
        erakunai.extend(l1)

kunai = sorted(erakunai, key=lambda x:(x[0],x[11],x[10],x[5])) 

print("kunai", len(kunai))

for i in reversed(range(0, len(kunai))):
    if (kunai[i][14] > 0) and (kunai[i][14] < 900): 
            perai = kunai.pop(i)

wbt = Workbook()
wst = wbt.active
j = 2
#derai.append(serai[0]) 
for i in range(0, len(kunai) - 1):
    if abs(kunai[i + 1][5] - kunai[i][5]) < 5 and abs(kunai[i + 1][6] - kunai[i][6]) < 5 and \
       (kunai[i + 1][0] == kunai[i][0]) and \
       (kunai[i + 1][11] == kunai[i][11]) and (kunai[i + 1][10] == kunai[i][10]):
        temperai.append(kunai[i])
        temperai.append(kunai[i + 1])
    else:
        if abs(kunai[i + 1][5] - kunai[i][5]) < 16 and abs(kunai[i + 1][6] - kunai[i][6]) < 4 and \
         (kunai[i + 1][0] == kunai[i][0]) and \
         (kunai[i + 1][11] == kunai[i][11]) and (kunai[i + 1][10] == kunai[i][10]) and \
         ((kunai[i + 1][9] / kunai[i][9] < 0.9) or (kunai[i][9] / kunai[i + 1][9] < 0.9)):
            temperai.append(kunai[i])
            temperai.append(kunai[i + 1])
        else:
            if len(temperai) > 0:
                temperai.append(kunai[i])
                strai = sorted(temperai, key=lambda x:(x[0],x[4]),reverse=True)
                mxrai = sorted(temperai, key=lambda x:(x[0],x[8]),reverse=True)
                mirai = sorted(temperai, key=lambda x:(x[0],x[7]))
#                       fn:0         ft:1        char:2      cls:3       rate:4      tx:5        ty:6     
#                       le:7         re:8        aw:9          CC:10      lc:11
                tElm = [strai[0][0],strai[0][1],strai[0][2],strai[0][3],strai[0][4],strai[0][5],strai[0][6],
                        mirai[0][7],mxrai[0][8],strai[0][9],strai[0][10],strai[0][11],strai[0][12],strai[0][13],strai[0][14]]
                derai.append(tElm)

                for i in range(0,len(temperai)):
                    j = j + i
                    wst.cell(row=j, column=2).value = temperai[i][0]
                    wst.cell(row=j, column=3).value = temperai[i][1]
                    wst.cell(row=j, column=4).value = temperai[i][2]
                    wst.cell(row=j, column=5).value = temperai[i][3]
                    wst.cell(row=j, column=6).value = temperai[i][4]
                    wst.cell(row=j, column=7).value = temperai[i][5]
                    wst.cell(row=j, column=8).value = temperai[i][6]
                    wst.cell(row=j, column=9).value = temperai[i][7]
                    wst.cell(row=j, column=10).value = temperai[i][8]
                    wst.cell(row=j, column=11).value = temperai[i][9]
                    wst.cell(row=j, column=12).value = temperai[i][10]
                    wst.cell(row=j, column=13).value = temperai[i][11]
#  wst.cell(row=j, column=14).value = serai[i][12]
                wst.cell(row=j, column=22).value = tElm[0]
                wst.cell(row=j, column=23).value = tElm[1]
                wst.cell(row=j, column=24).value = tElm[2]
                wst.cell(row=j, column=25).value = tElm[3]
                wst.cell(row=j, column=26).value = tElm[4]
                wst.cell(row=j, column=27).value = tElm[5]
                wst.cell(row=j, column=28).value = tElm[6]
                wst.cell(row=j, column=29).value = tElm[7]
                wst.cell(row=j, column=30).value = tElm[8]
                wst.cell(row=j, column=31).value = tElm[9]
                wst.cell(row=j, column=32).value = tElm[10]
                wst.cell(row=j, column=33).value = tElm[11]
                temperai = []
            else:
                derai.append(kunai[i])

wbt.save(dirOfPredx + "exNamevacktemperai" + ".xlsx") 



sdp = sorted(derai, key=lambda x:(x[0],x[11],x[5]))

print("len-sdp",len(sdp))

eraifilesv = ""
erailcls1sv = 0
erailcls2sv = 0
xerai = []

#          fn:0       ft:1     char:2     cls:3    rate:4     tx:5       ty:6     le:7       re:8     aw:9     CC:10      lc:11   cc2:12 
eraElm = [sdp[0][0],sdp[0][1],sdp[0][2],sdp[0][3],sdp[0][4],sdp[0][5],sdp[0][6],sdp[0][7],sdp[0][8],sdp[0][9],sdp[0][10],sdp[0][11],1]
#print(eraElm)
xerai.append(eraElm)
colc = 1
for i in range(1,len(sdp)):
    eraifile = sdp[i][0]
    erailcls1 = sdp[i][11]
    erailcls2 = sdp[i][10]
    npos = sdp[i-1][8]
    ppos = sdp[i][7]
    nupos = sdp[i-1][6]
    cppos = sdp[i][6]
    if eraifilesv != eraifile or erailcls1sv != erailcls1:
        npos = sdp[i][7]
        nupos = sdp[i][6]
        eraifilesv = eraifile
        erailcls1sv = erailcls1
        colc = 1
    if erailcls2sv != erailcls2:
        if abs(ppos - npos) < 6 and abs(cppos - nupos) < 6:
#            print("connect",eraifile,erailcls1,erailcls2sv,"-",erailcls2)
            colc = colc
            erailcls2sv = erailcls2
        else:
            npos = sdp[i][7]
            nupos = sdp[i][6]
            erailcls2sv = erailcls2
            colc = colc + 1
    diffx = abs(ppos - npos)
    diffy = abs(cppos - nupos)
    if diffy > 30:
            colc = colc + 1
    elif diffx > 14:
            colc = colc + 1
#               fn:0     ft:1     char:2     cls:3    rate:4     tx:5       ty:6     le:7      
#               re:8     aw:9     CC:10      lc:11   cc2:12 
    eraEsp = [sdp[i][0],sdp[i][1],sdp[i][2],sdp[i][3],sdp[i][4],sdp[i][5],sdp[i][6],sdp[i][7],
              sdp[i][8],sdp[i][9],sdp[i][10],sdp[i][11],colc]
    xerai.append(eraEsp)

sspt = sorted(xerai, key=lambda x:(x[0],x[11],x[5])) 

#print(xerai)  

wbt = Workbook()
wst = wbt.active
for i in range(0,len(xerai)):
  j = i + 2
  wst.cell(row=j, column=2).value = xerai[i][0]
  wst.cell(row=j, column=3).value = xerai[i][1]
  wst.cell(row=j, column=4).value = xerai[i][2]
  wst.cell(row=j, column=5).value = xerai[i][3]
  wst.cell(row=j, column=6).value = xerai[i][4]
  wst.cell(row=j, column=7).value = xerai[i][5]
  wst.cell(row=j, column=8).value = xerai[i][6]
  wst.cell(row=j, column=9).value = xerai[i][7]
  wst.cell(row=j, column=10).value = xerai[i][8]
  wst.cell(row=j, column=11).value = xerai[i][9]
  wst.cell(row=j, column=12).value = xerai[i][10]
  wst.cell(row=j, column=13).value = xerai[i][11]
  wst.cell(row=j, column=14).value = xerai[i][12]
wbt.save(dirOfPredx + "exNameerai" + ".xlsx") 

exOutNmsv = ""
rowidysv = 0
colidxsv = 0
charclxsv = 0
charclysv = 0
ratesv = 0
writeCnt = 0
charstr = ""

for clsElm in xerai:  
    exOutNm = clsElm[0]
    if exOutNmsv != exOutNm:
        if exOutNmsv != "":
            wb.save(dirOfPredx + exOutNmsv + ".xlsx")    
        wb = Workbook()
        ws = wb.active          
        exOutNmsv = exOutNm

        column_width = {'A': 2, 'B': 0.4, 'C': 0.4, 'D': 0.4, 'E': 0.4,'F': 0.4, 'G': 0.4, 'H': 0.4,
                    'I': 0.4, 'J': 0.4, 'K': 0.4, 'L': 0.4, 'M': 0.4,'N': 0.4, 'O': 0.4, 'P': 0.4,
                    'Q': 0.4, 'R': 0.4, 'S': 0.4, 'T': 0.4, 'U': 0.4,'V': 0.4, 'W': 0.4, 'X': 0.4,
                    'Y': 0.4, 'Z': 0.4, 'AA': 0.4, 'AB': 0.4, 'AC': 0.4,'AD': 0.4, 'AE': 0.4, 'AF': 0.4,
                    'AG': 0.4, 'AH': 0.4, 'AI': 0.4, 'AJ': 0.4, 'AK': 0.4,'AL': 0.4, 'AM': 0.4, 'AN': 0.4,
                    'AO': 0.4, 'AP': 0.4, 'AQ': 0.4, 'AR': 0.4, 'AS': 0.4,'AT': 0.4, 'AU': 0.4, 'AV': 0.4,
                    'AW': 0.4, 'AX': 0.4, 'AY': 0.4, 'AZ': 0.4,
                    'BA': 0.4, 'BB': 0.4, 'BC': 0.4, 'BD': 0.4,'BE': 0.4, 'BF': 0.4,
                    'BG': 0.4, 'BH': 0.4, 'BI': 0.4, 'BJ': 0.4, 'BK': 0.4,'BL': 0.4, 'BM': 0.4, 'BN': 0.4,
                    'BO': 0.4, 'BP': 0.4, 'BQ': 0.4, 'BR': 0.4, 'BS': 0.4,'BT': 0.4, 'BU': 0.4, 'BV': 0.4,
                    'BW': 0.4, 'BX': 0.4, 'BY': 0.4, 'BZ': 0.4,
                    'CA': 0.4, 'CB': 0.4, 'CC': 0.4, 'CD': 0.4,'CE': 0.4, 'CF': 0.4,
                    'CG': 0.4, 'CH': 0.4, 'CI': 0.4, 'CJ': 0.4, 'CK': 0.4,'CL': 0.4, 'CM': 0.4, 'CN': 0.4,
                    'CO': 0.4, 'CP': 0.4, 'CQ': 0.4, 'CR': 0.4, 'CS': 0.4,'CT': 0.4, 'CU': 0.4, 'CV': 0.4,
                    'CW': 0.4, 'CX': 0.4, 'CY': 0.4, 'CZ': 0.4,
                    'DA': 0.4, 'DB': 0.4, 'DC': 0.4, 'DD': 0.4,'DE': 0.4, 'DF': 0.4,
                    'DG': 0.4, 'DH': 0.4, 'DI': 0.4, 'DJ': 0.4, 'DK': 0.4,'DL': 0.4, 'DM': 0.4, 'DN': 0.4,
                    'DO': 0.4, 'DP': 0.4, 'DQ': 0.4, 'DR': 0.4, 'DS': 0.4,'DT': 0.4, 'DU': 0.4, 'DV': 0.4,
                    'DW': 0.4, 'DX': 0.4, 'DY': 0.4, 'DZ': 0.4,  
                    'EA': 0.4, 'EB': 0.4, 'EC': 0.4, 'ED': 0.4,'EE': 0.4, 'EF': 0.4,
                    'EG': 0.4, 'EH': 0.4, 'EI': 0.4, 'EJ': 0.4, 'EK': 0.4,'EL': 0.4, 'EM': 0.4, 'EN': 0.4,
                    'EO': 0.4, 'EP': 0.4, 'EQ': 0.4, 'ER': 0.4, 'ES': 0.4,'ET': 0.4, 'EU': 0.4, 'EV': 0.4,
                    'EW': 0.4, 'EX': 0.4, 'EY': 0.4, 'EZ': 0.4,
                    'FA': 0.4, 'FB': 0.4, 'FC': 0.4, 'FD': 0.4,'FE': 0.4, 'FF': 0.4,
                    'FG': 0.4, 'FH': 0.4, 'FI': 0.4, 'FJ': 0.4, 'FK': 0.4,'FL': 0.4, 'FM': 0.4, 'FN': 0.4,
                    'FO': 0.4, 'FP': 0.4, 'FQ': 0.4, 'FR': 0.4, 'FS': 0.4,'FT': 0.4, 'FU': 0.4, 'FV': 0.4,
                    'FW': 0.4, 'FX': 0.4, 'FY': 0.4, 'FZ': 0.4,
                    'GA': 0.4, 'GB': 0.4, 'GC': 0.4, 'GD': 0.4,'GE': 0.4, 'GF': 0.4,
                    'GG': 0.4, 'GH': 0.4, 'GI': 0.4, 'GJ': 0.4, 'GK': 0.4,'GL': 0.4, 'GM': 0.4, 'GN': 0.4,
                    'GO': 0.4, 'GP': 0.4, 'GQ': 0.4, 'GR': 0.4, 'GS': 0.4,'GT': 0.4, 'GU': 0.4, 'GV': 0.4,
                    'GW': 0.4, 'GX': 0.4, 'GY': 0.4, 'GZ': 0.4,
                    'HA': 0.4, 'HB': 0.4, 'HC': 0.4, 'HD': 0.4,'HE': 0.4, 'HF': 0.4,
                    'HG': 0.4, 'HH': 0.4, 'HI': 0.4, 'HJ': 0.4, 'HK': 0.4,'HL': 0.4, 'HM': 0.4, 'HN': 0.4,
                    'HO': 0.4, 'HP': 0.4, 'HQ': 0.4, 'HR': 0.4, 'HS': 0.4,'HT': 0.4, 'HU': 0.4, 'HV': 0.4,
                    'HW': 0.4, 'HX': 0.4, 'HY': 0.4, 'HZ': 0.4,
                    'IA': 0.4, 'IB': 0.4, 'IC': 0.4, 'ID': 0.4,'IE': 0.4, 'IF': 0.4,
                    'IG': 0.4, 'IH': 0.4, 'II': 0.4, 'IJ': 0.4, 'IK': 0.4,'IL': 0.4, 'IM': 0.4, 'IN': 0.4,
                    'IO': 0.4, 'IP': 0.4, 'IQ': 0.4, 'IR': 0.4, 'IS': 0.4,'IT': 0.4, 'IU': 0.4, 'IV': 0.4,
                    'IW': 0.4, 'IX': 0.4, 'IY': 0.4, 'IZ': 0.4,
                    'JA': 0.4, 'JB': 0.4, 'JC': 0.4, 'JD': 0.4,'JE': 0.4, 'JF': 0.4,
                    'JG': 0.4, 'JH': 0.4, 'JI': 0.4, 'JJ': 0.4, 'JK': 0.4,'JL': 0.4, 'JM': 0.4, 'JN': 0.4,
                    'JO': 0.4, 'JP': 0.4, 'JQ': 0.4, 'JR': 0.4, 'JS': 0.4,'JT': 0.4, 'JU': 0.4, 'JV': 0.4,
                    'JW': 0.4, 'JX': 0.4, 'JY': 0.4, 'JZ': 0.4,
                    'KA': 0.4, 'KB': 0.4, 'KC': 0.4, 'KD': 0.4,'KE': 0.4, 'KF': 0.4,
                    'KG': 0.4, 'KH': 0.4, 'KI': 0.4, 'KJ': 0.4, 'KK': 0.4,'KL': 0.4, 'KM': 0.4, 'KN': 0.4,
                    'KO': 0.4, 'KP': 0.4, 'KQ': 0.4, 'KR': 0.4, 'KS': 0.4,'KT': 0.4, 'KU': 0.4, 'KV': 0.4,
                    'KW': 0.4, 'KX': 0.4, 'KY': 0.4, 'KZ': 0.4,
                    'LA': 0.4, 'LB': 0.4, 'LC': 0.4, 'LD': 0.4,'LE': 0.4, 'LF': 0.4,
                    'LG': 0.4, 'LH': 0.4, 'LI': 0.4, 'LJ': 0.4, 'LK': 0.4,'LL': 0.4, 'LM': 0.4, 'LN': 0.4,
                    'LO': 0.4, 'LP': 0.4, 'LQ': 0.4, 'LR': 0.4, 'LS': 0.4,'LT': 0.4, 'LU': 0.4, 'LV': 0.4,
                    'LW': 0.4, 'LX': 0.4, 'LY': 0.4, 'LZ': 0.4,
                    'MA': 0.4, 'MB': 0.4, 'MC': 0.4, 'MD': 0.4,'ME': 0.4, 'MF': 0.4,
                    'MG': 0.4, 'MH': 0.4, 'MI': 0.4, 'MJ': 0.4, 'MK': 0.4,'ML': 0.4, 'MM': 0.4, 'MN': 0.4,
                    'MO': 0.4, 'MP': 0.4, 'MQ': 0.4, 'MR': 0.4, 'MS': 0.4,'MT': 0.4, 'MU': 0.4, 'MV': 0.4,
                    'MW': 0.4, 'MX': 0.4, 'MY': 0.4, 'MZ': 0.4,
                    'NA': 0.4, 'NB': 0.4, 'NC': 0.4, 'ND': 0.4,'NE': 0.4, 'NF': 0.4,
                    'NG': 0.4, 'NH': 0.4, 'NI': 0.4, 'NJ': 0.4, 'NK': 0.4,'NL': 0.4, 'NM': 0.4, 'NN': 0.4,
                    'NO': 0.4, 'NP': 0.4, 'NQ': 0.4, 'NR': 0.4, 'NS': 0.4,'NT': 0.4, 'NU': 0.4, 'NV': 0.4,
                    'NW': 0.4, 'NX': 0.4, 'NY': 0.4, 'NZ': 0.4,
                    'OA': 0.4, 'OB': 0.4, 'OC': 0.4, 'OD': 0.4,'OE': 0.4, 'OF': 0.4,
                    'OG': 0.4, 'OH': 0.4, 'OI': 0.4, 'OJ': 0.4, 'OK': 0.4,'OL': 0.4, 'OM': 0.4, 'ON': 0.4,
                    'OO': 0.4, 'OP': 0.4, 'OQ': 0.4, 'OR': 0.4, 'OS': 0.4,'OT': 0.4, 'OU': 0.4, 'OV': 0.4,
                    'OW': 0.4, 'OX': 0.4, 'OY': 0.4, 'OZ': 0.4,
                    'PA': 0.4, 'PB': 0.4, 'PC': 0.4, 'PD': 0.4,'PE': 0.4, 'PF': 0.4,
                    'PG': 0.4, 'PH': 0.4, 'PI': 0.4, 'PJ': 0.4, 'PK': 0.4,'PL': 0.4, 'PM': 0.4, 'PN': 0.4,
                    'PO': 0.4, 'PP': 0.4, 'PQ': 0.4, 'PR': 0.4, 'PS': 0.4,'PT': 0.4, 'PU': 0.4, 'PV': 0.4,
                    'PW': 0.4, 'PX': 0.4, 'PY': 0.4, 'PZ': 0.4,
                    'QA': 0.4, 'QB': 0.4, 'QC': 0.4, 'QD': 0.4,'QE': 0.4, 'QF': 0.4,
                    'QG': 0.4, 'QH': 0.4, 'QI': 0.4, 'QJ': 0.4, 'QK': 0.4,'QL': 0.4, 'QM': 0.4, 'QN': 0.4,
                    'QO': 0.4, 'QP': 0.4, 'QQ': 0.4, 'QR': 0.4, 'QS': 0.4,'QT': 0.4, 'QU': 0.4, 'QV': 0.4,
                    'QW': 0.4, 'QX': 0.4, 'QY': 0.4, 'QZ': 0.4,
                    'RA': 0.4, 'RB': 0.4, 'RC': 0.4, 'RD': 0.4,'RE': 0.4, 'RF': 0.4,
                    'RG': 0.4, 'RH': 0.4, 'RI': 0.4, 'RJ': 0.4, 'RK': 0.4,'RL': 0.4, 'RM': 0.4, 'RN': 0.4,
                    'RO': 0.4, 'RP': 0.4, 'RQ': 0.4, 'RR': 0.4, 'RS': 0.4,'RT': 0.4, 'RU': 0.4, 'RV': 0.4,
                    'RW': 0.4, 'RX': 0.4, 'RY': 0.4, 'RZ': 0.4,
                    'SA': 0.4, 'SB': 0.4, 'SC': 0.4, 'SD': 0.4,'SE': 0.4, 'SF': 0.4,
                    'SG': 0.4, 'SH': 0.4, 'SI': 0.4, 'SJ': 0.4, 'SK': 0.4,'SL': 0.4, 'SM': 0.4, 'SN': 0.4,
                    'SO': 0.4, 'SP': 0.4, 'SQ': 0.4, 'SR': 0.4, 'SS': 0.4,'ST': 0.4, 'SU': 0.4, 'SV': 0.4,
                    'SW': 0.4, 'SX': 0.4, 'SY': 0.4, 'SZ': 0.4,
                    'TA': 0.4, 'TB': 0.4, 'TC': 0.4, 'TD': 0.4,'TE': 0.4, 'TF': 0.4,
                    'TG': 0.4, 'TH': 0.4, 'TI': 0.4, 'TJ': 0.4, 'TK': 0.4,'TL': 0.4, 'TM': 0.4, 'TN': 0.4,
                    'TO': 0.4, 'TP': 0.4, 'TQ': 0.4, 'TR': 0.4, 'TS': 0.4,'TT': 0.4, 'TU': 0.4, 'TV': 0.4,
                    'TW': 0.4, 'TX': 0.4, 'TY': 0.4, 'TZ': 0.4}


 
        for col, width in column_width.items():
            ws.column_dimensions[col].width = width

        for shi in range(2,200):
            ws.row_dimensions[shi].height = 2
        keysDict = list(column_width.keys())

    if charclysv != clsElm[11]:
        if len(charstr) > 0: 
            ws.cell(row=rowidy, column=colidx).value = charstr
            ws.cell(row=rowidy, column=1).value = "a"
            ws.cell(row=rowidy, column=colidx).alignment = Alignment(vertical='center')  
            dicidx = keysDict[colidx + 1]
            ws.column_dimensions[dicidx].width = 1
            ws.row_dimensions[rowidy].height = 25
        charstr = ""
        charclysv = clsElm[11]
        charclxsv = clsElm[12]
        colidx = int(int(clsElm[5] / 10)) + 1
    else:
        if charclxsv != clsElm[12]:
            if len(charstr) > 0: 
                ws.cell(row=rowidy, column=colidx).value = charstr
                ws.cell(row=rowidy, column=1).value = "a"
                if (2 - column < 1):
                    wkcolumn = 1
                else:
                    wkcolumn = column - 2
                ws.cell(row=rowidy, column=wkcolumn).value = "!"
                ws.cell(row=rowidy, column=colidx).alignment = Alignment(vertical='center')
                dicidx = keysDict[colidx + 1]
                ws.column_dimensions[dicidx].width = 1
                ws.row_dimensions[rowidy].height = 25
            charstr = ""
            charclxsv = clsElm[12]
            colidx = int(int(clsElm[5] / 10)) + 1
    
    #rowidy = int(clsElm[11]) * 2 + 1
    rowidy = int(int(clsElm[6] / 45)) + 1

    if str(clsElm[2]) == "MinusMinus":
        char = "-"
    else:
        if str(clsElm[2]) == "DotDot":
            char = "!"
        else:
            char = str(clsElm[2])
    charstr = charstr + char

wb.save(dirOfPredx + exOutNmsv + ".xlsx")      